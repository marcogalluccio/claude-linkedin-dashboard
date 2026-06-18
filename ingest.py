#!/usr/bin/env python3
"""
LinkedIn Dashboard — ingest engine.

Merges the latest LinkedIn .xlsx export (quantitative) with the latest
qualitative/*.md (qualitative) and regenerates data.js, which dashboard.html reads.

Usage:  python3 ingest.py
        python3 ingest.py --xlsx data/2026-06-04-linkedin-export.xlsx --qual qualitative/2026-06-04-qualitative.md

Join key between the two layers is the LinkedIn post URL. Deterministic: same inputs -> same data.js.
"""
import openpyxl, datetime, json, re, sys, glob, os, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
KNOWN_GROUPS = {"Builder","Thought","Event","Consulting","Personal","Stablecoin","Hub"}

def pdate(s):
    s = str(s).strip()
    d, m, y = s.split("/")
    return datetime.date(int(y), int(m), int(d))

def latest(pattern):
    files = sorted(glob.glob(os.path.join(HERE, pattern)))
    if not files:
        sys.exit(f"Nessun file per il pattern {pattern!r} in {HERE}")
    return files[-1]  # filenames are date-prefixed -> last = most recent

def load_qual(path):
    md = open(path, encoding="utf-8").read()
    m = re.search(r"```json\s*\n(.*?)\n```", md, re.S)
    if not m:
        sys.exit(f"Nessun blocco ```json``` trovato in {path}")
    data = json.loads(m.group(1))
    by_url = {p["url"]: p for p in data.get("posts", []) if p.get("url")}
    return by_url, data.get("insights", [])

def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # SCOPERTA: totals
    sc = {r[0]: r[1] for r in wb["SCOPERTA"].iter_rows(values_only=True) if r[0]}
    impressions = int(sc.get("Impressioni", 0))
    reached = int(sc.get("Utenti raggiunti", 0))
    # INTERESSE: daily impressions + interactions
    irows = [(pdate(r[0]), int(r[1] or 0), int(r[2] or 0))
             for r in wb["INTERESSE"].iter_rows(min_row=2, values_only=True) if r[0]]
    start = irows[0][0]
    daily_imp = [r[1] for r in irows]
    daily_inter = [r[2] for r in irows]
    # FOLLOWER: total + daily new
    wf = wb["FOLLOWER"]
    foll_total = int(list(wf.iter_rows(max_row=1, values_only=True))[0][1])
    frows = [(pdate(r[0]), int(r[1] or 0)) for r in wf.iter_rows(min_row=4, values_only=True) if r[0]]
    daily_follnew = [r[1] for r in frows]
    # POST PRINCIPALI: per-url impressions (cols E-G) + interactions (cols A-C)
    posts = {}  # url -> {date, imp, inter}
    for r in wb["POST PRINCIPALI"].iter_rows(min_row=4, values_only=True):
        if r[0]:
            posts.setdefault(r[0], {"date": pdate(r[1]), "imp": 0, "inter": 0})["inter"] = int(r[2] or 0)
        if r[4]:
            posts.setdefault(r[4], {"date": pdate(r[5]), "imp": 0, "inter": 0})["imp"] = int(r[6] or 0)
    # DATI DEMOGRAFICI
    demo = {}
    for cat, val, pct in wb["DATI DEMOGRAFICI"].iter_rows(min_row=2, values_only=True):
        if cat is None or not isinstance(pct, (int, float)):
            continue
        demo.setdefault(cat, []).append([val, round(pct * 100, 2)])
    def top(cat, n=5):
        return demo.get(cat, [])[:n]
    demographics = {
        "localita":   top("Località"),
        "anzianita":  top("Anzianità"),
        "dimensione": top("Dimensioni azienda"),
        "settore":    top("Settori"),
    }
    return dict(impressions=impressions, reached=reached, start=start,
                daily_imp=daily_imp, daily_inter=daily_inter, daily_follnew=daily_follnew,
                foll_total=foll_total, posts=posts, demographics=demographics)

def build(xlsx_path, qual_path):
    q_by_url, insights = load_qual(qual_path)
    x = parse_xlsx(xlsx_path)

    merged = []
    unmatched, dropped, unknown_groups = [], 0, set()
    for url, xp in x["posts"].items():
        if xp["imp"] == 0:   # post non in classifica impressioni (solo lista interazioni): fuori dal ranking, interazioni già nei totali
            dropped += 1
            continue
        q = q_by_url.get(url)
        if q:
            grp = q.get("group", "Builder")
            post = {"hook": q["hook"], "type": q.get("type", "Originale"),
                    "pillar": q.get("pillar", grp), "group": grp,
                    "reaz": q.get("reaz"), "comm": q.get("comm"),
                    "imp": xp["imp"], "inter": xp["inter"],
                    "date": xp["date"].isoformat()}
            if q.get("enr"):
                post["enr"] = q["enr"]
        else:
            grp = "Builder"
            post = {"hook": f"(da etichettare · {xp['date'].isoformat()})", "type": "Export",
                    "pillar": grp, "group": grp, "reaz": None, "comm": None,
                    "imp": xp["imp"], "inter": xp["inter"], "date": xp["date"].isoformat()}
            unmatched.append(xp["date"].isoformat())
        if post["group"] not in KNOWN_GROUPS:
            unknown_groups.add(post["group"])
        merged.append(post)

    merged.sort(key=lambda p: -p["imp"])
    for i, p in enumerate(merged):
        p["rank"] = i + 1
    # reorder keys so rank is first
    merged = [{"rank": p.pop("rank"), **p} for p in merged]

    foll_base = x["foll_total"] - sum(x["daily_follnew"])
    # save-rate record from enriched posts
    rec = max(((p["enr"]["salv"] / p["imp"] * 100, p["hook"]) for p in merged
               if p.get("enr") and p["imp"] and p["enr"].get("salv") is not None), default=(0, ""))
    kpis = {
        "impressions": x["impressions"], "reached": x["reached"],
        "interactions": sum(x["daily_inter"]),
        "followers": x["foll_total"], "follGainedYear": sum(x["daily_follnew"]),
        "follBase": foll_base,
        "saveRateRecord": {"value": round(rec[0], 2), "label": rec[1]},
    }
    s = x["start"]
    end = s + datetime.timedelta(days=len(x["daily_imp"]) - 1)
    fname_date = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(xlsx_path))
    export_date = fname_date.group(1) if fname_date else end.isoformat()
    meta = {
        "exportDate": export_date,
        "periodStart": s.isoformat(), "periodEnd": end.isoformat(),
        "postCount": len(merged),
        "xlsx": os.path.relpath(xlsx_path, HERE),
        "qualitative": os.path.relpath(qual_path, HERE),
    }
    DATA = {"meta": meta, "kpis": kpis,
            "daily": {"start": [s.year, s.month, s.day],
                      "imp": x["daily_imp"], "inter": x["daily_inter"], "follNew": x["daily_follnew"]},
            "posts": merged, "demographics": x["demographics"], "insights": insights}
    return DATA, dict(unmatched=unmatched, dropped=dropped, unknown_groups=unknown_groups)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx"); ap.add_argument("--qual")
    a = ap.parse_args()
    xlsx_path = os.path.join(HERE, a.xlsx) if a.xlsx else latest("data/*.xlsx")
    qual_path = os.path.join(HERE, a.qual) if a.qual else latest("qualitative/*.md")
    DATA, diag = build(xlsx_path, qual_path)
    body = "// GENERATO da ingest.py — non modificare a mano. Rigenera con: python3 ingest.py\n"
    body += "window.DATA = " + json.dumps(DATA, ensure_ascii=False, indent=1) + ";\n"
    open(os.path.join(HERE, "data.js"), "w", encoding="utf-8").write(body)
    print(f"✓ data.js scritto")
    print(f"  fonte quantitativa : {os.path.relpath(xlsx_path, HERE)}")
    print(f"  fonte qualitativa  : {os.path.relpath(qual_path, HERE)}")
    print(f"  post in classifica : {DATA['meta']['postCount']}")
    print(f"  KPI                : {DATA['kpis']['impressions']:,} impr · {DATA['kpis']['reached']:,} reach · "
          f"{DATA['kpis']['interactions']:,} interaz · {DATA['kpis']['followers']} follower")
    if diag["unmatched"]:
        print(f"  ⚠ post nell'export senza qualitativo (da etichettare): {diag['unmatched']}")
    if diag["dropped"]:
        print(f"  · {diag['dropped']} post a 0 impressioni esclusi dalla classifica (interazioni già nei totali)")
    if diag["unknown_groups"]:
        print(f"  ⚠ pillar senza colore in dashboard.html (GROUPS): {sorted(diag['unknown_groups'])}")

if __name__ == "__main__":
    main()
